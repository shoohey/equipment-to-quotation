"""Excel解析モジュール"""

import re
from openpyxl import load_workbook
from models.equipment import EquipmentItem, EquipmentList, load_product_catalog, lookup_product


def parse_excel(file_path):
    """Excelファイルから機器データを抽出する"""
    wb = load_workbook(file_path, data_only=True)
    catalog = load_product_catalog()
    items = []

    for sheet in wb.worksheets:
        items.extend(_parse_sheet(sheet, catalog))

    equipment = EquipmentList(
        project_name=_extract_project_name(wb) or "（Excel読込）",
        client_name="",
        items=items,
    )
    return equipment


def _parse_sheet(sheet, catalog):
    """シートから機器データを抽出"""
    items = []
    header_row = _find_header_row(sheet)

    if header_row is None:
        return _parse_by_model_search(sheet, catalog)

    col_map = _map_columns(sheet, header_row)
    if not col_map:
        return _parse_by_model_search(sheet, catalog)

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
        item = _extract_item_from_row(row, col_map, catalog)
        if item:
            items.append(item)

    return items


def _find_header_row(sheet):
    """ヘッダー行を検出"""
    keywords = ["型番", "品番", "型式", "メーカー", "品名", "数量"]
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        row_text = " ".join(str(c) for c in row if c)
        matches = sum(1 for kw in keywords if kw in row_text)
        if matches >= 2:
            return row_idx
    return None


def _map_columns(sheet, header_row):
    """列のマッピングを作成"""
    col_map = {}
    mapping = {
        "型番": "model", "品番": "model", "型式": "model",
        "メーカー": "maker", "製造元": "maker",
        "品名": "name", "機器名": "name", "名称": "name",
        "数量": "quantity", "台数": "quantity", "個数": "quantity",
        "単位": "unit",
        "定価": "list_price", "単価": "list_price",
        "設置場所": "location", "場所": "location", "備考": "location",
    }

    for cell in sheet[header_row]:
        if cell.value:
            val = str(cell.value).strip()
            for keyword, field in mapping.items():
                if keyword in val:
                    col_map[field] = cell.column - 1
                    break

    return col_map if "model" in col_map else {}


def _extract_item_from_row(row, col_map, catalog):
    """行からアイテムを抽出"""
    cells = [cell.value for cell in row]

    model_idx = col_map.get("model")
    model = str(cells[model_idx]).strip() if model_idx is not None and cells[model_idx] else ""
    if not model or model == "None":
        return None

    product = lookup_product(model, catalog)

    maker = ""
    if "maker" in col_map and cells[col_map["maker"]]:
        maker = str(cells[col_map["maker"]]).strip()
    elif product:
        maker = product.get("maker", "")

    name = ""
    if "name" in col_map and cells[col_map["name"]]:
        name = str(cells[col_map["name"]]).strip()
    elif product:
        name = product.get("name", "")

    quantity = 1
    if "quantity" in col_map and cells[col_map["quantity"]]:
        try:
            quantity = int(float(str(cells[col_map["quantity"]])))
        except (ValueError, TypeError):
            quantity = 1

    unit = "台"
    if "unit" in col_map and cells[col_map["unit"]]:
        unit = str(cells[col_map["unit"]]).strip()
    elif product:
        unit = product.get("unit", "台")

    list_price = 0
    if "list_price" in col_map and cells[col_map["list_price"]]:
        try:
            list_price = int(float(str(cells[col_map["list_price"]]).replace(",", "").replace("¥", "")))
        except (ValueError, TypeError):
            list_price = 0
    if list_price == 0 and product:
        list_price = product.get("list_price", 0)

    category = ""
    if product:
        category = product.get("category", "")

    location = ""
    if "location" in col_map and cells[col_map["location"]]:
        location = str(cells[col_map["location"]]).strip()

    return EquipmentItem(
        category=category,
        maker=maker,
        model=model,
        name=name or model,
        quantity=quantity,
        unit=unit,
        list_price=list_price,
        location=location,
    )


def _parse_by_model_search(sheet, catalog):
    """ヘッダーが見つからない場合、型番をセル全体から検索"""
    items = []
    all_models = {}
    for category, products in catalog.items():
        for p in products:
            all_models[p["model"]] = {**p, "category": category}

    found_models = {}
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip()
            for model, info in all_models.items():
                if model in cell_str and model not in found_models:
                    found_models[model] = info

    for model, info in found_models.items():
        items.append(EquipmentItem(
            category=info["category"],
            maker=info["maker"],
            model=model,
            name=info["name"],
            quantity=1,
            unit=info.get("unit", "台"),
            list_price=info["list_price"],
        ))

    return items


def _extract_project_name(wb):
    """ワークブックから件名を推定"""
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                if "工事" in cell or "ビル" in cell or "建築" in cell:
                    return cell.strip()
    return None
